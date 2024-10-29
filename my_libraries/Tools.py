from typing import Literal
import logging
import os.path

import json
import openpyxl
import colorama
import datetime as dt
import pytz
import psutil
import requests






class Files:

    def __init__(self):
        if not os.path.exists(f'files'): os.mkdir("files")

    def get_json(self, name: str) -> dict:
        file = f'files/{name}' if name[-5:] == '.json' else f'files/{name}.json'
        if os.path.exists(file):
            with open(file) as f:
                config = json.load(f)
                return config

    def get_excel(self, name: str) -> list:
        file = f'files/{name}' if name[-5:] == '.xlsx' else f'files/{name}.xlsx'
        if os.path.exists(file):
            excel_data = openpyxl.load_workbook(file)
            worksheet = excel_data.worksheets[0]
            columns = list(list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0])
            rows = worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True)
            settings = []
            for row in rows:
                row = list(row)
                coin = {}
                for n, column in enumerate(columns):
                    if column == None:
                        continue
                    coin[str(column)] = row[n]
                settings.append(coin)
            return settings

class Functions():

    def divide_list(self, lst, x):
        """Разделяем список на Х равных частей"""

        lst = [i for i in lst if i is not None]
        size = len(lst) // x
        leftovers = len(lst) % x
        result = []
        for i in range(x):
            start = i * size + min(i, leftovers)
            end = (i + 1) * size + min(i + 1, leftovers)
            result.append(lst[start:end])
        return result

def config() -> dict:
    with open('files/config.json') as f:
        data = json.load(f)
    return data

def timenow():
    return dt.datetime.utcnow().replace(microsecond=0)
